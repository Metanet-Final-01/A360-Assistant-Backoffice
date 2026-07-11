"""두 버전(agent_label) 비교를 AB_comparison_report.xlsx(a360-eval-sandbox)와 같은 스타일의
엑셀로 내보낸다 — Overview(집계 평균) + Per-Case Comparison(케이스별 비교) 두 시트.
색상·레이아웃 상수는 그 파일(build_ab_report.py)에서 그대로 옮겨왔다.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .log_schema import EvalRunRecord

# frontend/views/eval_results.py의 _FIXED_METRICS와 값을 맞춰뒀다 — 화면에서 본
# "개선사항 비교"와 여기서 내려받는 엑셀이 서로 다른 지표 집합을 보여주면 안 되므로.
# 프론트/백엔드가 별도 프로세스라 리스트를 import로 공유할 수 없어 값만 동일하게 유지한다.
FIXED_METRICS = {
    "pm4py_fitness", "pm4py_precision",
    "worfbench_precision", "worfbench_recall", "worfbench_f1_score",
}

NAVY = "1F2A44"
TEAL = "0E7C7B"
RED_FILL = "FBE4E4"
GREEN_FILL = "E4F5E9"
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color=NAVY, bold=True, size=16)
SUB_FONT = Font(color="55606E", size=10, italic=True)
THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_row(ws: Worksheet, row: int, ncols: int, fill: str = NAVY) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autosize(ws: Worksheet, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _delta_fill(delta: float | None) -> PatternFill | None:
    if delta is None:
        return None
    if delta > 0.0001:
        return PatternFill("solid", fgColor=GREEN_FILL)
    if delta < -0.0001:
        return PatternFill("solid", fgColor=RED_FILL)
    return None


def _metrics_of(r: EvalRunRecord) -> dict[str, float]:
    """EvalRunRecord.score는 채점 엔진마다 다른 지표를 대표값으로 복사해 넣은 별칭이라
    (예: pm4py 행은 score=fitness) 일부러 안 씀 — FIXED_METRICS로 이름이 고정된
    metrics만 비교 대상으로 쓴다(자세한 이유는 frontend/views/eval_results.py의
    _FIXED_METRICS 주석 참고)."""
    return {m.name: m.value for m in r.metrics if m.name in FIXED_METRICS}


def _group_by_case(runs: list[EvalRunRecord]) -> dict[str, dict[str, float]]:
    """case_id별로 묶어 지표를 평균 낸다. 같은 case_id로 여러 번 기록된 로그(예: 같은
    케이스를 재실행)가 있어도 마지막 것만 조용히 채택하지 않고 전부 반영한다."""
    sums: dict[str, dict[str, list[float]]] = {}
    for r in runs:
        bucket = sums.setdefault(r.case_id, {})
        for name, value in _metrics_of(r).items():
            bucket.setdefault(name, []).append(value)
    return {
        case_id: {name: sum(vals) / len(vals) for name, vals in metrics.items()}
        for case_id, metrics in sums.items()
    }


def build_comparison_xlsx(
    runs_a: list[EvalRunRecord], runs_b: list[EvalRunRecord], label_a: str, label_b: str
) -> bytes:
    by_case_a = _group_by_case(runs_a)
    by_case_b = _group_by_case(runs_b)
    common_cases = sorted(set(by_case_a) & set(by_case_b))
    metric_names = sorted(
        set().union(*(by_case_a[c].keys() & by_case_b[c].keys() for c in common_cases))
    ) if common_cases else []

    wb = Workbook()

    # ---------------- Sheet 1: Overview ----------------
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    ws["B2"] = f"{label_a} vs {label_b} — 평가 결과 비교"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"공통 case_id {len(common_cases)}개 기준 지표 평균 비교 (A360-Assistant-Ops에서 자동 생성)"
    ws["B3"].font = SUB_FONT
    ws.merge_cells("B2:F2")
    ws.merge_cells("B3:F3")

    r = 5
    ws.cell(row=r, column=2, value="집계 평균").font = Font(bold=True, size=12, color=NAVY)
    r += 1
    headers = ["지표", f"버전 A ({label_a})", f"버전 B ({label_b})", "delta (B - A)", "변화율"]
    for c, h in enumerate(headers, start=2):
        ws.cell(row=r, column=c, value=h)
    _style_header_row(ws, r, len(headers) + 1, fill=NAVY)
    r += 1

    def avg(metrics_by_case: dict[str, dict[str, float]], key: str) -> float | None:
        vals = [metrics_by_case[c][key] for c in common_cases if key in metrics_by_case[c]]
        return round(sum(vals) / len(vals), 4) if vals else None

    for name in metric_names:
        a_val, b_val = avg(by_case_a, name), avg(by_case_b, name)
        delta = round(b_val - a_val, 4) if a_val is not None and b_val is not None else None
        pct = f"{delta / a_val * 100:+.1f}%" if delta is not None and a_val else "n/a"
        ws.cell(row=r, column=2, value=name).border = BORDER
        ws.cell(row=r, column=3, value=a_val).border = BORDER
        ws.cell(row=r, column=4, value=b_val).border = BORDER
        dcell = ws.cell(row=r, column=5, value=delta)
        dcell.border = BORDER
        fill = _delta_fill(delta)
        if fill:
            dcell.fill = fill
        ws.cell(row=r, column=6, value=pct).border = BORDER
        r += 1

    _autosize(ws, {1: 2, 2: 26, 3: 22, 4: 22, 5: 16, 6: 12})

    # ---------------- Sheet 2: Per-Case Comparison ----------------
    ws2 = wb.create_sheet("Per-Case Comparison")
    ws2.sheet_view.showGridLines = False
    cols = ["case_id"]
    for name in metric_names:
        cols += [f"{name}_A", f"{name}_B", f"{name}_delta"]
    for c, h in enumerate(cols, start=1):
        ws2.cell(row=1, column=c, value=h)
    _style_header_row(ws2, 1, len(cols), fill=TEAL)

    row = 2
    for case_id in common_cases:
        ma, mb = by_case_a[case_id], by_case_b[case_id]
        ws2.cell(row=row, column=1, value=case_id).border = BORDER
        col = 2
        for name in metric_names:
            va, vb = ma.get(name), mb.get(name)
            delta = round(vb - va, 4) if va is not None and vb is not None else None
            ws2.cell(row=row, column=col, value=va).border = BORDER
            ws2.cell(row=row, column=col + 1, value=vb).border = BORDER
            dcell = ws2.cell(row=row, column=col + 2, value=delta)
            dcell.border = BORDER
            fill = _delta_fill(delta)
            if fill:
                dcell.fill = fill
            col += 3
        row += 1

    widths = {1: 34}
    for i in range(2, len(cols) + 1):
        widths[i] = 12
    _autosize(ws2, widths)
    ws2.freeze_panes = "B2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
