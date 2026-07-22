"""RAGAS chunk_size 그리드서치 결과를 엑셀 하나로 정리한다.

시트 6개: 요약 / 전체 / doc_page / action_schema 전체 / action_schema-jar /
action_schema-llm_agent. jar/llm_agent 분리는 원격 RAG 코퍼스 DB
(RAG_DATABASE_URL)의 rag_documents.metadata->>'schema_source'를 골드셋
case_id의 reference_doc_ids[0]로 직접 조회해서 판정한다(case_id는 5개
chunk_size 실험 전체에서 완전히 동일한 100개 집합임을 확인했음 — 재현
가능한 조인).

사용:
    python -m scripts.ragas_eval.chunk_size_report
"""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_REPO_ROOT = Path(__file__).resolve().parents[2]
_RAG_SERVER_ROOT = _REPO_ROOT / "rag-server"
_EVAL_RUNS_PATH = _REPO_ROOT / "ops-server" / "backend" / "data" / "eval_runs.jsonl"
_GOLDSET_PATH = (
    _REPO_ROOT / "ops-server" / "backend" / "app" / "eval" / "ragas_eval"
    / "cases" / "rag_goldset_v1.json"
)
_OUTPUT_PATH = (
    _REPO_ROOT / "docs" / "local" / "gpt_handoff_2026-07-20"
    / "RAGAS_CHUNK_SIZE_ALL_RESULTS_2026-07-20.xlsx"
)


def _resolve_output_path() -> Path:
    """대상 파일이 열려 있어 쓰기 권한이 없으면(엑셀에서 열어둔 경우 등)
    타임스탬프 붙인 대체 경로로 저장한다 — 결과를 못 남기고 그냥 실패하는 것보다
    낫다."""
    try:
        _OUTPUT_PATH.touch(exist_ok=True)
        return _OUTPUT_PATH
    except PermissionError:
        import datetime

        stamp = datetime.datetime.now().strftime("%H%M%S")
        return _OUTPUT_PATH.with_name(_OUTPUT_PATH.stem + f"_{stamp}" + _OUTPUT_PATH.suffix)

_METRICS = [
    "faithfulness", "answer_relevancy", "context_precision", "context_recall",
    "answer_correctness", "hit_at_1", "hit_at_3", "hit_at_5",
    "reciprocal_rank", "evidence_coverage",
]
_METRIC_LABELS = {
    "faithfulness": "Faithfulness", "answer_relevancy": "AnswerRelevancy",
    "context_precision": "ContextPrecision", "context_recall": "ContextRecall",
    "answer_correctness": "AnswerCorrectness", "hit_at_1": "Hit@1",
    "hit_at_3": "Hit@3", "hit_at_5": "Hit@5", "reciprocal_rank": "MRR",
    "evidence_coverage": "EvidenceCoverage",
}

_HEADER_FILL = PatternFill(start_color="1F6F8B", end_color="1F6F8B", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_WIN_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
_TITLE_FONT = Font(bold=True, size=14)
_NOTE_FONT = Font(italic=True, color="666666")


def _metric_map(rec: dict) -> dict[str, float]:
    return {m["name"]: m["value"] for m in rec.get("metrics", [])}


def _avg(vals: list[float | None]) -> float | None:
    vals = [v for v in vals if v is not None]
    return round(sum(vals) / len(vals), 4) if vals else None


# 글자 수 기준 chunk_size가 실제로 몇 토큰인지 — 전체 코퍼스(doc_page 3,258건)를
# 진짜로 chunk_text()에 통과시켜 나온 청크들의 평균 토큰 수(2026-07-20 실측,
# o200k_base=gpt-4o-mini/추정 프로덕션, cl100k_base=text-embedding-3-small).
# 새로 실험을 안 돌리고 기존 5개 결과에 참고용 라벨만 붙이는 용도.
_TOKEN_ESTIMATE_BY_CHUNK_SIZE = {
    300: {"o200k_avg": 128.4, "cl100k_avg": 186.0},
    600: {"o200k_avg": 245.1, "cl100k_avg": 354.8},
    900: {"o200k_avg": 343.1, "cl100k_avg": 496.7},
    1200: {"o200k_avg": 422.5, "cl100k_avg": 611.6},
    1500: {"o200k_avg": 491.1, "cl100k_avg": 710.7},
}


def load_chunk_experiment_records() -> dict[str, list[dict]]:
    """eval_runs.jsonl에서 chunk_size 실험 기록을 읽되, 지금 골드셋의 active
    케이스 집합에 속한 case_id만 남긴다 — 과거(2026-07-18, 100케이스 구표본)와
    오늘(2026-07-20, jar30+llm_agent30로 확장된 129케이스) 기록이 같은
    agent_label 밑에 섞여 있어서, active 집합으로 걸러야 지금 골드셋 기준
    결과만 뽑힌다.

    골드셋 JSON을 직접 읽는다(runner.py를 임포트하지 않음) — ops-server와
    rag-server 둘 다 최상위 패키지명이 'app'이라, 이 함수 이후에
    build_case_schema_source_map()이 rag-server의 app.rag.config를 임포트할 때
    이미 캐시된 ops-server의 app 패키지와 충돌해 조용히 깨지는 문제가 있었다
    (2026-07-20, source_documents.py에서 겪은 것과 같은 원인)."""
    goldset_cases = json.loads(_GOLDSET_PATH.read_text(encoding="utf-8"))
    active_case_ids = {
        c["case_id"] for c in goldset_cases
        if c.get("status") == "approved" and c.get("dataset_membership") == "active"
    }

    # (agent_label, case_id)로 중복 제거 — 원래 100케이스 표본의 대부분(69 doc_page +
    # 기존 action_schema 30건)은 지금도 active라서, 2026-07-18 구실행과 2026-07-20
    # 재실행 두 기록이 같은 case_id로 둘 다 남아있다. active 집합 필터만으론 이 중복을
    # 못 걸러서(228건 = 실제 129건의 거의 2배로 나온 버그, 2026-07-20 발견) 최신
    # logged_at 하나만 남긴다.
    latest_by_key: dict[tuple[str, str], dict] = {}
    with _EVAL_RUNS_PATH.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            if rec.get("source") != "ragas_chunk_experiment" or rec["case_id"] not in active_case_ids:
                continue
            key = (rec["agent_label"], rec["case_id"])
            existing = latest_by_key.get(key)
            if existing is None or rec.get("logged_at", "") > existing.get("logged_at", ""):
                latest_by_key[key] = rec

    recs_by_label: dict[str, list[dict]] = defaultdict(list)
    for (label, _case_id), rec in latest_by_key.items():
        recs_by_label[label].append(rec)
    return recs_by_label


def build_case_schema_source_map() -> dict[str, str]:
    """골드셋 각 케이스(action_schema만)가 jar/llm_agent 중 어디서 왔는지, 원격 RAG
    코퍼스 DB를 직접 조회해서 판정한다.

    ⚠️ `rag_documents.id`가 아니라 `parent_id`로 조인해야 한다 — 원격 테이블은
    이미 청크 단위로 쪼개져 저장되어 있어서(운영 chunk_size=1200), 문서 하나가
    여러 chunk_index로 나뉘면 `id`는 청크마다 다르고 원본 문서를 가리키는 건
    `parent_id`뿐이다. 처음에 `id`로 조인했다가 다중청크 문서 전부가 "고아
    문서"로 오판되는 버그가 있었다(2026-07-19 발견·수정) — 실제로는 로컬/원격
    전체 스캔 결과 진짜 고아는 골드셋 113건 중 2건뿐이었다."""
    sys.path.insert(0, str(_RAG_SERVER_ROOT))
    from app.rag import config  # noqa: E402
    import psycopg  # noqa: E402

    cases = json.loads(_GOLDSET_PATH.read_text(encoding="utf-8"))
    conn = psycopg.connect(config.database_dsn())
    cur = conn.cursor()

    case_to_source: dict[str, str] = {}
    for case in cases:
        doc_ids = case.get("reference_doc_ids") or []
        if not doc_ids:
            continue
        cur.execute(
            "select distinct source_type, metadata->>'schema_source' "
            "from rag_documents where parent_id = %s",
            (doc_ids[0],),
        )
        row = cur.fetchone()
        if row is None or row[0] != "action_schema":
            continue
        case_to_source[case["case_id"]] = row[1] or "unknown"
    conn.close()
    return case_to_source


def build_scope_table(
    recs_by_label: dict[str, list[dict]],
    labels: list[str],
    scope_filter,
) -> list[list]:
    """scope_filter(case_id) -> bool 인 케이스만 골라 chunk_size별 평균 지표 표를 만든다.
    맨 뒤 두 컬럼은 그 chunk_size(글자)가 실제로 평균 몇 토큰인지 참고값(전체 코퍼스
    실측, 새로 실험 안 돌림) — "글자 수 기준"이라는 실험 설계 자체를 바꾸는 게 아니라
    결과 해석을 돕는 라벨이다."""
    rows = [["chunk_size(글자)", "n", *[_METRIC_LABELS[m] for m in _METRICS],
             "≈토큰(o200k)", "≈토큰(cl100k,임베딩)"]]
    for label in labels:
        recs = [r for r in recs_by_label[label] if scope_filter(r["case_id"])]
        n = len(recs)
        chunk_size = int(label.replace("_ov0", "").replace("cs", ""))
        row = [label.replace("_ov0", ""), n]
        for m in _METRICS:
            row.append(_avg([_metric_map(r).get(m) for r in recs]))
        token_est = _TOKEN_ESTIMATE_BY_CHUNK_SIZE.get(chunk_size, {})
        row.append(token_est.get("o200k_avg"))
        row.append(token_est.get("cl100k_avg"))
        rows.append(row)
    return rows


def write_sheet(wb: openpyxl.Workbook, sheet_name: str, title: str, rows: list[list]) -> None:
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))

    header_row_idx = 3
    for col_idx, header in enumerate(rows[0], start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    data_rows = rows[1:]
    for r_offset, row in enumerate(data_rows):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=header_row_idx + 1 + r_offset, column=c_idx, value=value)

    # 지표 컬럼(3번째 컬럼부터)마다 최댓값 셀을 강조 — 전부 "높을수록 좋음" 지표라
    # 최댓값 = 그 지표에서 이긴 chunk_size. 맨 뒤 토큰 환산 2컬럼은 "좋고 나쁨"이
    # 없는 참고값이라 강조 대상에서 뺀다.
    n_metric_cols = len(rows[0]) - 2 - 2
    for metric_col_offset in range(n_metric_cols):
        col_idx = 3 + metric_col_offset
        values = []
        for r_offset in range(len(data_rows)):
            v = data_rows[r_offset][col_idx - 1]
            values.append(v if isinstance(v, (int, float)) else None)
        real_values = [v for v in values if v is not None]
        if not real_values:
            continue
        best = max(real_values)
        for r_offset, v in enumerate(values):
            if v == best:
                ws.cell(row=header_row_idx + 1 + r_offset, column=col_idx).fill = _WIN_FILL

    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)
    for col_idx in range(1, len(rows[0]) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16 if col_idx > 2 else 12


def write_summary_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("요약", 0)
    ws["A1"] = "RAGAS chunk_size 그리드서치 — 전체 결과 요약"
    ws["A1"].font = _TITLE_FONT

    lines = [
        "",
        "실험: 300/600/900/1200/1500 (overlap=0), 승인 골드셋 129케이스(active) × 5개 chunk_size,",
        "      gpt-4o-mini 채점 (2026-07-20, jar/llm_agent 각각 30건씩 채운 확장판)",
        "재현성: 5개 chunk_size 실험 전부 지금 골드셋의 active 129케이스 집합 기준으로 필터링함.",
        "",
        "각 시트 맨 뒤 두 컬럼(≈토큰)은 그 chunk_size(글자)가 실제 코퍼스에서 평균 몇 토큰인지",
        "참고값입니다 — 토큰 기준으로 실험을 다시 돌린 게 아니라, 기존 글자 기준 결과에 환산",
        "라벨만 붙인 것입니다(o200k_base=생성/추정 프로덕션, cl100k_base=임베딩).",
        "",
        "핵심 결론",
        "1) 단일 chunk_size로는 모든 문서유형을 동시에 만족 못함 — cs600은 전 구간 탈락.",
        "2) doc_page(원본 문서, 산문): cs300이 압도적 — 작게 쪼갤수록 검색 정밀도가 좋음.",
        "3) action_schema(정형 텍스트, 총 60케이스, jar 30 + llm_agent 30)는 cs1200~1500 쪽으로",
        "   기웁니다 — ContextRecall/EvidenceCoverage/Faithfulness가 cs1200 이상에서 최고.",
        "4) doc_page와 action_schema(llm_agent)는 같은 원본 문서에서 나올 수 있지만 서로",
        "   다른 별개 문서다 — LLM 파싱이 원본을 대체하지 않고 구조화된 사본을 추가로 만든다",
        "   (docs.jsonl 원본 3,258건 그대로 보존 확인, cmd_export_for_agent는 읽기 전용).",
        "",
        "권고: 문서유형별(source_type) 차등 chunk_size 적용은 데이터 근거 확실(doc_page=작게,",
        "      action_schema=크게). 상세 설계는 RAG_DYNAMIC_CHUNK_SIZE_INVESTIGATION_2026-07-19.md 참고.",
        "",
        "시트 구성: 전체 / doc_page / action_schema(전체) / action_schema-jar / action_schema-llm_agent",
    ]
    for i, line in enumerate(lines, start=2):
        cell = ws.cell(row=i, column=1, value=line)
        if line.startswith(("핵심", "권고", "시트")):
            cell.font = Font(bold=True)
        elif line and not line[0].isdigit() and "실험" not in line and "재현" not in line:
            pass
    ws.column_dimensions["A"].width = 100


def main() -> None:
    recs_by_label = load_chunk_experiment_records()
    labels = sorted(recs_by_label.keys(), key=lambda l: int(l.replace("cs", "").split("_")[0]))

    print("jar/llm_agent 판정을 위해 원격 RAG 코퍼스 DB 조회 중...")
    case_schema_source = build_case_schema_source_map()
    print(f"  action_schema 케이스 {len(case_schema_source)}건 판정 완료")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb)

    all_rows = build_scope_table(recs_by_label, labels, lambda cid: True)
    write_sheet(wb, "전체", "전체(ALL, n=100/사이즈)", all_rows)

    def is_doc_page(case_id: str, recs_by_label=recs_by_label, labels=labels) -> bool:
        for label in labels:
            for r in recs_by_label[label]:
                if r["case_id"] == case_id:
                    return (r.get("raw") or {}).get("source_type") == "doc_page"
        return False

    def is_action_schema(case_id: str, recs_by_label=recs_by_label, labels=labels) -> bool:
        for label in labels:
            for r in recs_by_label[label]:
                if r["case_id"] == case_id:
                    return (r.get("raw") or {}).get("source_type") == "action_schema"
        return False

    doc_page_rows = build_scope_table(recs_by_label, labels, is_doc_page)
    write_sheet(wb, "doc_page", "doc_page 문서(원본 산문)", doc_page_rows)

    action_schema_rows = build_scope_table(recs_by_label, labels, is_action_schema)
    write_sheet(wb, "action_schema 전체", "action_schema 전체(jar+llm_agent)", action_schema_rows)

    jar_rows = build_scope_table(
        recs_by_label, labels, lambda cid: case_schema_source.get(cid) == "jar"
    )
    write_sheet(wb, "action_schema-jar", "action_schema · schema_source=jar", jar_rows)

    llm_rows = build_scope_table(
        recs_by_label, labels, lambda cid: case_schema_source.get(cid) == "llm_agent"
    )
    write_sheet(wb, "action_schema-llm_agent", "action_schema · schema_source=llm_agent", llm_rows)

    _OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    out_path = _resolve_output_path()
    wb.save(out_path)
    print(f"저장 완료: {out_path}")


if __name__ == "__main__":
    main()
