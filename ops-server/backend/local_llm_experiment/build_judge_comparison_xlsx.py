"""로컬모델(EXAONE-4.5/4.0, reasoning 끔)과 gpt-4o-mini를 RAGAS judge로 썼을 때
결과를 모아 엑셀로 정리한다. reasoning-ON(4.5/4.0 둘 다) 조건은 실측으로 실패
확인됐고 완료 데이터가 없어서 별도 시트에 실패 근거만 기록한다.
"""
import json
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment

EVAL_RUNS = Path("data/eval_runs.jsonl")
OUTPUT_PATH = Path(
    r"c:/Users/KDH/Documents/VisualStudio Code/A360-Assistant/A360-Assistant-Ops/docs/local/gpt_handoff_2026-07-22/"
    r"RAGAS_JUDGE_MODEL_COMPARISON_2026-07-22.xlsx"
)

CONDITIONS = [
    ("cs1200_judge_local_noreason", "judge=EXAONE4.5끔 (생성=로컬4.5)"),
    ("cs1200_exaone40_nr_fixedgen", "judge=EXAONE4.0끔 (생성=로컬4.5,동일재사용)"),
    ("cs1200_judge_gpt4omini", "judge=gpt-4o-mini (생성=로컬4.5)"),
    ("cs1200_gen_gpt4omini", "judge=gpt-4o-mini (생성=gpt-4o-mini)"),
]

METRIC_KEYS = ["faithfulness", "answer_relevancy", "context_precision", "context_recall",
               "answer_correctness", "hit_at_1", "hit_at_3", "hit_at_5", "reciprocal_rank",
               "evidence_coverage"]
METRIC_LABELS = {
    "faithfulness": "Faithfulness", "answer_relevancy": "AnswerRelevancy",
    "context_precision": "ContextPrecision", "context_recall": "ContextRecall",
    "answer_correctness": "AnswerCorrectness", "hit_at_1": "Hit@1", "hit_at_3": "Hit@3",
    "hit_at_5": "Hit@5", "reciprocal_rank": "MRR", "evidence_coverage": "EvidenceCoverage",
}


def load_latest_by_case(label: str) -> dict:
    by_eval_id: dict[str, list] = defaultdict(list)
    with EVAL_RUNS.open(encoding="utf-8") as f:
        for line in f:
            rec = json.loads(line)
            if rec.get("agent_label") == label:
                by_eval_id[rec["evaluation_id"]].append(rec)
    if not by_eval_id:
        return {}
    latest_eval_id = max(by_eval_id.keys(), key=lambda e: int(e.rsplit("_", 1)[1]))
    return {r["case_id"]: r for r in by_eval_id[latest_eval_id]}


def metric_map(rec: dict) -> dict:
    return {m["name"]: m["value"] for m in rec.get("metrics", [])}


def avg_metrics(recs_by_case: dict) -> dict:
    sums: dict[str, list] = defaultdict(list)
    for rec in recs_by_case.values():
        for k, v in metric_map(rec).items():
            if v is not None:
                sums[k].append(v)
    return {k: sum(v) / len(v) for k, v in sums.items() if v}


def write_sheet(wb, name, rows):
    ws = wb.create_sheet(name)
    for row in rows:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(10, min(60, max_len + 2))
    return ws


def main() -> None:
    condition_data = {label: load_latest_by_case(agent_label) for agent_label, label in CONDITIONS}
    for agent_label, label in CONDITIONS:
        print(f"{label}: {len(condition_data[label])}건")

    common_case_ids = None
    for label, recs in condition_data.items():
        ids = set(recs.keys())
        common_case_ids = ids if common_case_ids is None else (common_case_ids & ids)
    common_case_ids = sorted(common_case_ids)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 요약 시트
    summary_rows = [["조건", "케이스 수"] + [METRIC_LABELS[k] for k in METRIC_KEYS]]
    for agent_label, label in CONDITIONS:
        recs = condition_data[label]
        m = avg_metrics(recs)
        row = [label, len(recs)] + [round(m.get(k), 4) if m.get(k) is not None else None for k in METRIC_KEYS]
        summary_rows.append(row)
    ws_summary = write_sheet(wb, "요약", summary_rows)
    wb.move_sheet("요약", offset=-len(wb.sheetnames))

    # 케이스별 상세 시트 (조건 3개 나란히)
    detail_header = ["case_id", "question_type", "source_type"]
    for _, label in CONDITIONS:
        for k in METRIC_KEYS:
            detail_header.append(f"{label}_{METRIC_LABELS[k]}")
    detail_rows = [detail_header]
    for cid in common_case_ids:
        first_rec = condition_data[CONDITIONS[0][1]][cid]
        raw = first_rec.get("raw", {})
        row = [cid, raw.get("question_type"), raw.get("source_type")]
        for _, label in CONDITIONS:
            m = metric_map(condition_data[label][cid])
            row += [round(m.get(k), 4) if m.get(k) is not None else None for k in METRIC_KEYS]
        detail_rows.append(row)
    write_sheet(wb, "케이스별_상세", detail_rows)

    # 답변 원문 비교 시트
    answer_header = ["case_id", "question", "ground_truth"] + [f"{label}_answer" for _, label in CONDITIONS]
    answer_rows = [answer_header]
    for cid in common_case_ids:
        first_raw = condition_data[CONDITIONS[0][1]][cid]["raw"]
        row = [cid, first_raw.get("question"), first_raw.get("ground_truth")]
        for _, label in CONDITIONS:
            row.append(condition_data[label][cid]["raw"].get("answer"))
        answer_rows.append(row)
    ws_answers = write_sheet(wb, "답변원문", answer_rows)
    for row_cells in ws_answers.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws_answers.column_dimensions[col_letter].width = 50

    # reasoning-ON 실패 근거 시트
    notes_rows = [
        ["조건", "결과", "근거"],
        [
            "로컬-EXAONE4.5(추론켬)",
            "완료 데이터 없음(제외)",
            "10건 정식 실행 중 57분+ 경과하도록 진행 없어 강제 종료. 이전 1건 스모크 테스트에서 "
            "reasoning_content가 max_tokens 전부 소모해 최종 답변을 못 낸 것 실측 확인(1+1 질문 기준 "
            "reasoning_len 716자, content는 빈 값에 가까움).",
        ],
        [
            "로컬-EXAONE4.0(추론켬)",
            "완료 데이터 없음(제외)",
            "2026-07-22에 2건 스모크 테스트 실행 — 10개 채점 작업(2케이스x5지표) 중 6개 실패"
            "(LLMDidNotFinishException 5, TimeoutError 1) = 실패율 60%, 소요시간 8.8분/2건"
            "(케이스당 4~5분, 다른 조건 대비 수십 배 느림). 단순 프롬프트('1+1은?')에서는 성공했지만"
            "실제 RAGAS 채점 프롬프트(검색문서 포함, 훨씬 김)에서는 자주 실패해 전체 10건 실행은 진행하지 않음.",
        ],
    ]
    write_sheet(wb, "추론켬_제외사유", notes_rows)

    summary_ws = wb["요약"]
    summary_ws.insert_rows(1, amount=6)
    summary_ws["A1"] = "RAGAS Judge 모델 비교 (2026-07-22)"
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws["A2"] = "검색 테이블: rag_documents_eval_cs1200_ov0 (top_k=5, vector-only, reranker 없음), 10케이스 고정"
    summary_ws["A3"] = "1~3행: 생성모델을 로컬-EXAONE4.5로 고정하고 judge만 바꿔서 비교(공정 비교). 4행: 생성까지 gpt-4o-mini로 바꾼 조건(참고용)"
    summary_ws["A4"] = "로컬 judge는 reasoning을 꺼야만(enable_thinking:false) 안정적으로 동작함 확인됨 (아래부터 데이터)"
    summary_ws["A5"] = "reasoning-ON 조건(4.5/4.0 둘 다)은 '추론켬_제외사유' 시트 참고 — 완료 데이터 없어 이 표에서 제외"
    summary_ws["A6"] = ""

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print("saved:", OUTPUT_PATH)


if __name__ == "__main__":
    main()
